import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def encontrar_carpeta_descripciones_bbdd_mas_reciente(ruta_base):
    """
    Busca en 'ruta_base' todas las subcarpetas que cumplan el patrón:
      ^YYYYMMDD - Descripciones BBDD.*$
    (Por ejemplo: '20250305 - Descripciones BBDD' o
                  '20250305 - Descripciones BBDD + GTE')
    Devuelve la ruta a la de fecha mayor, o None si no hay candidatas.
    """
    carpetas_candidatas = []
    patron = r'^(\d{8}) - Descripciones BBDD.*$'
    for nombre in os.listdir(ruta_base):
        ruta_completa = os.path.join(ruta_base, nombre)
        if os.path.isdir(ruta_completa):
            match = re.match(patron, nombre)
            if match:
                fecha_str = match.group(1)  # YYYYMMDD
                carpetas_candidatas.append((int(fecha_str), ruta_completa))

    if not carpetas_candidatas:
        return None
    # Tomar la de fecha más alta
    carpeta_mas_reciente = max(carpetas_candidatas, key=lambda x: x[0])
    return carpeta_mas_reciente[1]

def localizar_carpeta_final_y_fecha(carpeta_base):
    """
    Dada una carpeta cuyo nombre cumple '^YYYYMMDD - Descripciones BBDD.*$':
    - Extraemos la fecha (YYYYMMDD).
    - Si no tiene texto adicional tras 'BBDD', esa carpeta es la "carpeta final".
    - Si sí tiene texto adicional, buscamos dentro subcarpeta 'BD'.

    Devolvemos (carpeta_final, fecha_str).
    Si algo falla, devolvemos (None, None).
    """
    nombre_carpeta = os.path.basename(carpeta_base)
    patron = r'^(\d{8}) - Descripciones BBDD(.*)$'
    match = re.match(patron, nombre_carpeta)
    if not match:
        return (None, None)

    fecha_str = match.group(1)  # YYYYMMDD
    resto = match.group(2).strip()  # lo que sobra tras 'BBDD'

    if not resto:
        # No hay texto extra => la carpeta base es la final
        return (carpeta_base, fecha_str)
    else:
        # Hay texto extra => buscar subcarpeta 'BD'
        ruta_bd = os.path.join(carpeta_base, "BD")
        if os.path.isdir(ruta_bd):
            return (ruta_bd, fecha_str)
        else:
            return (None, None)

def buscar_archivo_rev_gramatical(ruta_extraccion):
    """
    Busca un XLSX que termine con '_rev_gramatical.xlsx' en la carpeta dada.
    """
    for fichero in os.listdir(ruta_extraccion):
        if fichero.lower().endswith('_rev_gramatical.xlsx'):
            return os.path.join(ruta_extraccion, fichero)
    return None

def ajustar_anchos_columna_openpyxl(worksheet):
    """
    Ajustar anchuras de columna para ~135 px en la primera y ~685 px en las demás.
    """
    from openpyxl.utils import get_column_letter
    pix_a_excel = lambda px: px / 7.0

    worksheet.column_dimensions[get_column_letter(1)].width = round(pix_a_excel(135), 1)
    for col_idx in [2,3,4]:
        worksheet.column_dimensions[get_column_letter(col_idx)].width = round(pix_a_excel(685), 1)

def guardar_como_xlsx(df, ruta_salida):
    from openpyxl import Workbook
    with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        hoja = writer.sheets['Sheet1']
        ajustar_anchos_columna_openpyxl(hoja)

def main():
    # Ajustar según tu entorno
    # ruta_base = r"\\s02-ean\DataAcquisition\TRANSLATIONS\01_Descriptions & Resources"
    ruta_base = r"C:\Users\Saul\Desktop\s02-ean\DataAcquisition\TRANSLATIONS\01_Descriptions & Resources"

    # 1) Encuentra la carpeta con fecha + "Descripciones BBDD"
    carpeta_encontrada = encontrar_carpeta_descripciones_bbdd_mas_reciente(ruta_base)
    if not carpeta_encontrada:
        print("No se encontró ninguna carpeta 'YYYYMMDD - Descripciones BBDD'. Saliendo...")
        return

    # 2) Determinar si la carpeta es final o si hay subcarpeta "BD"
    carpeta_final, fecha_carpeta = localizar_carpeta_final_y_fecha(carpeta_encontrada)
    if not carpeta_final:
        print("No se pudo localizar la carpeta final ni extraer la fecha. Saliendo...")
        return

    # 3) Ubicamos la carpeta 'Extracción'
    carpeta_extraccion = os.path.join(carpeta_final, "Extracción")
    if not os.path.isdir(carpeta_extraccion):
        print(f"No existe la carpeta 'Extracción' en {carpeta_final}. Saliendo...")
        return

    # 4) Buscar XLSX que termine en '_rev_gramatical.xlsx'
    archivo_rev = buscar_archivo_rev_gramatical(carpeta_extraccion)
    if not archivo_rev:
        print("No se encontró ningún '_rev_gramatical.xlsx'. Saliendo...")
        return

    # 5) Leer DataFrame, comprobar columnas
    df_original = pd.read_excel(archivo_rev)
    columnas_esperadas = {"DIC_NUM_ID", "ESP", "ENG"}
    if not columnas_esperadas.issubset(df_original.columns):
        print("El XLSX no contiene las columnas requeridas (DIC_NUM_ID, ESP, ENG). Saliendo...")
        return

    # 6) Idiomas a procesar
    idiomas = [
        "CAT", "DEU", "DNK", "ELL", "FIN", "FRA", "HUN", "ITA", "KOR",
        "NLD", "NNO", "PLN", "PRT", "RON", "SWE", "TUR"
    ]

    # 7) Crear carpeta "Enviados"
    carpeta_enviados = os.path.join(carpeta_final, "Enviados")
    os.makedirs(carpeta_enviados, exist_ok=True)

    # 8) Generar archivos para cada idioma
    for iso_code in idiomas:
        df_nuevo = df_original.copy()
        df_nuevo[iso_code] = ""  # Columna en blanco

        filename_base = f"{fecha_carpeta}_Descriptions_{iso_code}"
        if iso_code in ["ELL","KOR","TUR"]:
            # XLSX
            ruta_xlsx = os.path.join(carpeta_enviados, filename_base + ".xlsx")
            guardar_como_xlsx(df_nuevo, ruta_xlsx)
        else:
            # CSV
            ruta_csv = os.path.join(carpeta_enviados, filename_base + ".csv")
            df_nuevo.to_csv(ruta_csv, sep=';', encoding='utf-8', index=False)

    print(f"¡Proceso completado! Archivos creados en '{carpeta_enviados}'")

if __name__ == "__main__":
    main()
