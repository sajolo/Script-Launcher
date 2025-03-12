import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from datetime import datetime
import mysql.connector

# Directorio donde se almacenan las banderas
FLAGS_DIR = r"C:\Users\Saul\Desktop\s02-ean\DataAcquisition\Script-Launcher\images\images\flags"

# Configuración de la base de datos MySQL
db_config = {
    'user': 'smarquez',
    'password': 'helvetia',
    'host': 'localhost',
    'database': 'prc_db',
}

# Lista de países
countries_iso = [
    "DEU", "ESP", "POR", "FRA", "ITA", "GBR", "CHE", "NLD", "GRC", "PLN", "NOR", 
    "KOR", "AUT", "CZE", "FIN", "ROU", "DNK", "BEL", "HUN", "IRL", 
    "SVK", "SWE", "IND", "ZAF", "TUN", "NAM"
]

# Lista de marcas
brands = [
    "ALFA ROMEO", "AUDI", "BMW", "CITROËN", "DAEWOO/CHEVROLET", "DACIA", "DS", "FIAT", "FORD",
    "HONDA", "HYUNDAI", "IVECO", "JAGUAR", "JEEP", "KIA", "LANCIA", "LAND ROVER", "LEXUS",
    "MAN", "MAZDA", "MERCEDES BENZ", "MG", "MINI", "MITSUBISHI", "NISSAN", "OPEL",
    "PEUGEOT", "PORSCHE", "RENAULT", "RENAULT TRUCKS", "ROVER/MG", "SAAB", "SEAT", "SKODA", "SMART", "SUBARU", "SUZUKI",
    "TATA", "TESLA", "TOYOTA", "VOLVO", "VOLVO TRUCKS", "VOLKSWAGEN"
]

def get_latest_effective_date(country_iso, brand_name):
    connection = mysql.connector.connect(**db_config)
    cursor = connection.cursor()
    query = (
        "SELECT MAX(last_file_date) FROM tariffs_dates "
        "WHERE country_iso = %s AND brand_name = %s"
    )
    cursor.execute(query, (country_iso, brand_name))
    result = cursor.fetchone()[0]
    cursor.fetchall()  # Limpiar resultados no leídos
    cursor.close()
    connection.close()
    if result and result <= datetime.now().date():
        return result.strftime("%Y-%m-%d")
    return None

def update_price_file_report_ext(country_iso, brand_desc, effective_date):
    """
    Actualiza la tabla price_file_report_ext. 
    Las columnas PRICE_FILE_SOURCE y CONVERTED_FROM_DIFFERENT_SOURCE se dejan en blanco 
    (o se pueden obtener si lo deseas).
    """
    connection = mysql.connector.connect(**db_config)
    cursor = connection.cursor()
    query = (
        "REPLACE INTO price_file_report_ext (COUNTRY_ISO, BRAND_DESC, EFFECTIVE_DATE, PRICE_FILE_SOURCE, CONVERTED_FROM_DIFFERENT_SOURCE) "
        "VALUES (%s, %s, %s, '', '')"
    )
    cursor.execute(query, (country_iso, brand_desc, effective_date if effective_date else None))
    connection.commit()
    cursor.close()
    connection.close()

def create_or_update_xlsx():
    print("Generando informe de tarifas...")

    # Eliminar cualquier archivo anterior que comience por "TARIFFS_REPORT_"
    output_dir = r"C:\Users\Saul\Desktop\s02-ean\DataAcquisition\TARIFAS_ORIGINALES\_DESCARGAS"
    for filename in os.listdir(output_dir):
        if filename.startswith("TARIFFS_REPORT_"):
            os.remove(os.path.join(output_dir, filename))

    # Crear un nuevo libro
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tariffs_Report"

    # Ajustar el ancho de columna A a 21.5
    sheet.column_dimensions["A"].width = 21.5  

    # Los países van de la columna B a 11.5
    for col_idx in range(2, 2 + len(countries_iso)):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = 11.5

    # Insertar las marcas en filas desde la 3 en adelante (orden alfabético)
    sorted_brands = sorted(brands)
    for i, marca in enumerate(sorted_brands, start=3):
        sheet.cell(row=i, column=1, value=marca)

    # Insertar países en fila 2 (desde columna B en adelante) con sus banderas, 
    # añadiendo 6 espacios para que la bandera no tape el ISO
    sorted_countries = sorted(countries_iso)
    for j, c_iso in enumerate(sorted_countries, start=2):
        cell_val = "      " + c_iso  # 6 espacios 
        sheet.cell(row=2, column=j, value=cell_val)
        # Insertar bandera
        flag_path = os.path.join(FLAGS_DIR, f"{c_iso}.png")
        if os.path.exists(flag_path):
            img = Image(flag_path)
            img.width, img.height = 20, 12
            col_letter = openpyxl.utils.get_column_letter(j)
            sheet.add_image(img, f"{col_letter}2")

    # Ahora, en cada intersección (marca, país) se inserta la EFFECTIVE_DATE 
    # centrada horizontalmente
    for i, marca in enumerate(sorted_brands, start=3):
        for j, c_iso in enumerate(sorted_countries, start=2):
            effective_date = get_latest_effective_date(c_iso, marca)
            # Actualizar la BBDD
            update_price_file_report_ext(c_iso, marca, effective_date)
            if effective_date:
                cell = sheet.cell(row=i, column=j, value=effective_date)
                # Centrar la fecha horizontalmente
                cell.alignment = Alignment(horizontal="center")
            else:
                # Dejarla en blanco y centrar en blanco
                cell = sheet.cell(row=i, column=j, value="")
                cell.alignment = Alignment(horizontal="center")

    # Generar el nombre final
    report_filename = f"TARIFFS_REPORT_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    report_path = os.path.join(output_dir, report_filename)

    # Guardar el archivo
    workbook.save(report_path)
    print("Informe generado, cierre esta ventana para finalizar...")
    input()

if __name__ == "__main__":
    create_or_update_xlsx()
